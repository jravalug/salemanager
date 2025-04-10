from collections import defaultdict
from datetime import datetime
from io import BytesIO
import json
from flask import (
    Blueprint,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import Workbook

from dateutil.relativedelta import relativedelta

from app.forms import MonthForm
from app.models import Business, Sale, SaleDetail
from app.services import SalesReportService, BusinessService
from app.utils import generate_excel_sales_by_date, generate_excel_ipv


bp = Blueprint("report", __name__, url_prefix="/business/<int:business_id>/report")

sales_service = SalesReportService()
business_service = BusinessService()


@bp.route("/monthly-sales-by-product", methods=["GET", "POST"])
def monthly_sales_by_produc(business_id):
    # Obtener el negocio (asegúrate de que exista)
    business = Business.query.get_or_404(business_id)
    business_filters = business_service.get_parent_filters(business)

    # Inicializar variables
    daily_sales = []
    selected_month = None
    excluded_sales = []
    excluded_products = []

    if request.method == "POST":
        # Obtener el mes seleccionado desde el formulario
        selected_month = request.form.get("month")

        if not selected_month:
            flash("Por favor, selecciona un mes.", "error")
            return render_template(
                "report/monthly_sales_by_product.html",
                business=business,
                daily_sales=daily_sales,
                selected_month=selected_month,
            )

        try:

            sales, formated_daily_sales = sales_service.get_daily_sales(
                selected_month,
                business_filters["business_id"],
                business_filters.get("specific_business_id"),
            )

            # Procesar los datos
            sales_by_day = {}
            for sale in sales:
                date_key = sale.date.strftime("%Y-%m-%d")
                if date_key not in sales_by_day:
                    sales_by_day[date_key] = {
                        "total_products": 0,
                        "total_income": 0,
                        "products": {},
                    }

                # Ordenar los productos dentro de la venta por nombre
                sorted_products = sorted(
                    sale.products, key=lambda sale_detail: sale_detail.product.name
                )
                # logging.debug("Datos generados para el reporte: %s", sale.products)

                for sale_detail in sorted_products:
                    product_key = sale_detail.product.name
                    if product_key not in sales_by_day[date_key]["products"]:
                        sales_by_day[date_key]["products"][product_key] = {
                            "quantity": 0,
                            "total_amount": 0,
                            "orders": set(),  # Usamos un conjunto para evitar duplicados,
                        }

                    sales_by_day[date_key]["products"][product_key][
                        "quantity"
                    ] += sale_detail.quantity
                    sales_by_day[date_key]["products"][product_key][
                        "total_amount"
                    ] += sale_detail.total_price

                    # Añadir la tupla (sale_id, sale_number) al conjunto
                    sales_by_day[date_key]["products"][product_key]["orders"].add(
                        (sale.id, sale.sale_number)
                    )

                # Actualizar totales
                sales_by_day[date_key]["total_products"] += sum(
                    sale_detail.quantity for sale_detail in sale.products
                )
                sales_by_day[date_key]["total_income"] += sum(
                    sale_detail.total_price for sale_detail in sale.products
                )

            # Formatear los datos para la plantilla
            for date, data in sales_by_day.items():
                # Ordenar los productos por nombre
                sorted_products = sorted(
                    data["products"].items(),
                    key=lambda item: item[
                        0
                    ],  # Ordenar por el nombre del producto (item[0])
                )

                daily_sales.append(
                    {
                        "date": date,
                        "total_products": data["total_products"],
                        "total_income": data["total_income"],
                        "products": [
                            {
                                "name": name,
                                "quantity": product["quantity"],
                                "total_amount": product["total_amount"],
                                "orders": sorted(
                                    list(product["orders"]), key=lambda order: order[1]
                                ),
                            }
                            for name, product in sorted_products  # Usar los productos ordenados
                        ],
                    }
                )

        except ValueError:
            # Manejar errores si el formato del mes es incorrecto
            flash("Por favor, selecciona un mes válido.", "error")

    # logging.basicConfig(level=logging.DEBUG)
    # logging.debug("Datos generados para el reporte: %s", daily_sales)

    return render_template(
        "report/monthly_sales_by_product.html",
        business=business,
        daily_sales=daily_sales,
        daily_sales_json=json.dumps(daily_sales),
        selected_month=selected_month,
    )


@bp.route("/monthly-sales-by-date", methods=["GET", "POST"])
def monthly_sales_by_date(business_id):
    """Endpoint para obtener las ventas mensuales agrupadas por día."""

    #     # Obtener el negocio
    business = Business.query.get_or_404(business_id)
    business_filters = business_service.get_parent_filters(business)

    form = MonthForm()

    # Inicializar variables

    daily_sales = []
    formated_daily_sales = []
    selected_month = None

    def redirect_to_monthly_sales_by_date():
        return redirect(url_for("report.monthly_sales_by_date", business=business))

    if form.validate_on_submit():
        month_str = form.month.data

        try:
            daily_sales, formated_daily_sales = sales_service.get_daily_sales(
                month_str,
                business_filters["business_id"],
                business_filters.get("specific_business_id"),
            )

        except ValueError as e:
            flash(f"Error en el formato del mes: {e}", "error")
            redirect_to_monthly_sales_by_date()

        except Exception as e:
            flash("Ocurrió un error inesperado.", "error")
            redirect_to_monthly_sales_by_date()

        month_totals = sales_service.get_monthly_totals(daily_sales)

        return render_template(
            "report/monthly_sales_by_date.html",
            business=business,
            form=form,
            daily_sales=formated_daily_sales,
            daily_sales_json=json.dumps(formated_daily_sales),
            month_totals=month_totals,
            selected_month=month_str,
        )

    return render_template(
        "report/monthly_sales_by_date.html",
        business=business,
        form=form,
        daily_sales=[],
        selected_month=selected_month,
    )


@bp.route("/export-to-excel/sales-by-date", methods=["POST"])
def export_to_excel_sales_by_day(business_id):
    # Obtener el negocio
    business = Business.query.get_or_404(business_id)
    # Recuperar los datos del reporte mensual
    month = request.form.get("selected_month")
    daily_sales_json = request.form.get("daily_sales_export")
    if not daily_sales_json:
        flash("No hay datos disponibles para exportar.", "error")
        return redirect(
            url_for("report.monthly_sales_by_date", business_id=business.id)
        )
    try:
        # Convertir los datos JSON a una estructura de Python
        data = json.loads(daily_sales_json)
    except json.JSONDecodeError:
        flash("Los datos recibidos no son válidos.", "error")
        return redirect(
            url_for("report.monthly_sales_by_date", business_id=business.id)
        )

    try:
        # Generar el reporte en MS Excel
        excel_file = generate_excel_sales_by_date(business, data, month)
    except Exception as e:
        flash("Ocurrio un error al tratar de generar el reporte en MS Excel.", "error")
        return redirect(
            url_for("report.monthly_sales_by_date", business_id=business.id)
        )

    # Enviar el archivo como respuesta
    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"RMV_{business.name}_{month}.xlsx",
    )


@bp.route("/export-to-excel/sales-by-product", methods=["POST"])
def export_to_excel_sales_by_product(business_id):
    # Obtener el negocio
    business = Business.query.get_or_404(business_id)

    try:
        # Recuperar los datos del reporte mensual
        data = json.loads(request.form.get("daily_sales_export", "[]"))
        month = request.form.get("selected_month", "_")
        print(f"En export_to_excel_sales_by_product:{month}")
        if not data:
            flash("No hay datos disponibles para exportar.", "error")
            return redirect(
                url_for("report.monthly_sales_by_produc", business_id=business_id)
            )
    except json.JSONDecodeError:
        flash("Los datos recibidos no son válidos.", "error")
        return redirect(
            url_for("report.monthly_sales_by_produc", business_id=business_id)
        )

    print("Datos generados para el reporte: %s", data)

    # Procesar los datos para agrupar por producto
    product_summary = defaultdict(
        lambda: {"quantity": 0, "total_amount": 0, "orders": defaultdict(set)}
    )

    for day in data:
        if "date" not in day or "products" not in day:
            flash("Los datos recibidos no tienen el formato esperado.", "error")
            return redirect(
                url_for("report.monthly_sales_by_produc", business_id=business_id)
            )

        # Extraer el día de la fecha
        date_str = day["date"]
        try:
            day_number = datetime.strptime(date_str, "%Y-%m-%d").strftime(
                "%d"
            )  # Formato "01", "02", etc.
        except ValueError:
            flash("El formato de fecha en los datos no es válido.", "error")
            return redirect(
                url_for("report.monthly_sales_by_produc", business_id=business_id)
            )

        for product in day["products"]:
            if (
                "name" not in product
                or "quantity" not in product
                or "total_amount" not in product
                or "orders" not in product
            ):
                flash("Los datos recibidos no tienen el formato esperado.", "error")
                return redirect(
                    url_for("report.monthly_sales_by_produc", business_id=business_id)
                )

            product_name = product["name"]
            product_summary[product_name]["quantity"] += product["quantity"]
            product_summary[product_name]["total_amount"] += product["total_amount"]

            # Agregar los números de orden al conjunto correspondiente al día
            for order in product["orders"]:
                product_summary[product_name]["orders"][day_number].add(order[1])

    # Crear un archivo Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Mensual"

    # Encabezados
    headers = ["PRODUCTO", "CANTIDAD", "IMPORTE", "ORDENES"]
    sheet.append(headers)

    # Rellenar el archivo con los datos
    for product_name, details in product_summary.items():
        # Formatear las órdenes en el nuevo formato "01[1,2,3], 02[4,5], ..."
        orders_formatted = []
        for day_number, orders_set in details["orders"].items():
            orders_list = sorted(list(orders_set))
            orders_formatted.append(f"{day_number}[{','.join(map(str, orders_list))}]")
        orders_column = ", ".join(orders_formatted)

        row = [
            product_name,
            details["quantity"],
            details["total_amount"],
            orders_column,  # Columna ORDENES con el nuevo formato
        ]
        sheet.append(row)

    # Guardar el archivo en memoria
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    # Enviar el archivo como respuesta
    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"RMVxProductos_{business.name}_{month}.xlsx",
    )


@bp.route("/export-to-excel/sales-by-product-by-date", methods=["POST"])
def export_to_excel_sales_by_product_by_date(business_id):
    # Obtener el negocio
    business = Business.query.get_or_404(business_id)

    try:
        # Recuperar los datos del reporte mensual
        data = json.loads(request.form.get("daily_sales_export", "[]"))
        month = request.form.get("selected_month", "_")
        print(f"En export_to_excel_sales_by_product_by_date:{month}")
        if not data:
            flash("No hay datos disponibles para exportar.", "error")
            return redirect(
                url_for("report.monthly_sales_by_produc", business_id=business_id)
            )
    except json.JSONDecodeError:
        flash("Los datos recibidos no son válidos.", "error")
        return redirect(
            url_for("report.monthly_sales_by_produc", business_id=business_id)
        )

    # Procesar los datos para organizarlos por día
    sales_by_day = defaultdict(list)

    for day in data:
        if "date" not in day or "products" not in day:
            flash("Los datos recibidos no tienen el formato esperado.", "error")
            return redirect(
                url_for("report.monthly_sales_by_produc", business_id=business_id)
            )

        date_str = day["date"]
        try:
            formatted_date = datetime.strptime(date_str, "%Y-%m-%d").strftime(
                "%d/%m/%y"
            )  # Formato "DD/MM/YY"
        except ValueError:
            flash("El formato de fecha en los datos no es válido.", "error")
            return redirect(
                url_for("report.monthly_sales_by_produc", business_id=business_id)
            )

        for product in day["products"]:
            if (
                "name" not in product
                or "quantity" not in product
                or "total_amount" not in product
            ):
                flash("Los datos recibidos no tienen el formato esperado.", "error")
                return redirect(
                    url_for("report.monthly_sales_by_produc", business_id=business_id)
                )

            product_name = product["name"]
            quantity = product["quantity"]
            price_per_unit = product["total_amount"] / quantity if quantity > 0 else 0
            total_amount = product["total_amount"]

            sales_by_day[formatted_date].append(
                {
                    "product": product_name,
                    "quantity": quantity,
                    "price_per_unit": price_per_unit,
                    "total_amount": total_amount,
                }
            )

    # Crear un archivo Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Mensual"

    # Rellenar el archivo con los datos
    row_offset = 1  # Para manejar múltiples días
    for date, products in sales_by_day.items():
        # Agregar la cabecera del día
        sheet.cell(row=row_offset, column=1, value="DIA")
        sheet.cell(row=row_offset, column=2, value=date)
        row_offset += 1

        # Agregar los encabezados de las columnas
        headers = ["PRODUCTO", "CANTIDAD", "PRECIO UNITARIO", "IMPORTE"]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=row_offset, column=col_idx, value=header)
        row_offset += 1

        # Agregar los productos del día
        for product in products:
            sheet.cell(row=row_offset, column=1, value=product["product"])
            sheet.cell(row=row_offset, column=2, value=product["quantity"])
            sheet.cell(
                row=row_offset, column=3, value=round(product["price_per_unit"], 2)
            )
            sheet.cell(
                row=row_offset, column=4, value=round(product["total_amount"], 2)
            )
            row_offset += 1

        # Agregar una fila en blanco entre días
        row_offset += 1

    # Guardar el archivo en memoria
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    # Enviar el archivo como respuesta
    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"RMVxProductos_Diario_{business.name}_{month}.xlsx",
    )


@bp.route("/ipv-report", methods=["GET", "POST"])
def ipv_report(business_id):
    business = Business.query.get_or_404(business_id)
    business_filters = business_service.get_parent_filters(business)

    daily_sales = []
    selected_month = None

    if request.method == "POST":
        month_str = request.form.get("month")
        if not month_str:
            flash("Selecciona un mes válido.", "error")
            return redirect(url_for("report.ipv_report", business_id=business_id))

        try:
            selected_month = datetime.strptime(month_str, "%Y-%m").date()
            start_date = selected_month.replace(day=1)
            end_date = start_date + relativedelta(months=1, days=-1)

            # Consultar ventas del mes
            if business.is_general:
                sales = (
                    Sale.query.filter(
                        Sale.business_id == business_filters["business_id"],
                        Sale.date.between(start_date, end_date),
                    )
                    .join(Sale.products)  # Relación entre Sale y SaleDetail
                    .join(SaleDetail.product)  # Relación entre SaleDetail y Product
                    .order_by(Sale.date.asc())
                    .all()
                )
            else:
                sales = (
                    Sale.query.filter(
                        Sale.business_id == business_filters["business_id"],
                        Sale.specific_business_id
                        == business_filters["specific_business_id"],
                        Sale.date.between(start_date, end_date),
                    )
                    .join(Sale.products)  # Relación entre Sale y SaleDetail
                    .join(SaleDetail.product)  # Relación entre SaleDetail y Product
                    .order_by(Sale.date.asc())
                    .all()
                )
            print(f"sales: {sales}")

            # Obtener todos los productos non_food involucrados en las ventas del mes
            non_food_products = {}
            for sale in sales:
                for sale_product in sale.products:
                    product = sale_product.product
                    if product.category not in [
                        "comida",
                        "postre",
                        "trago",
                        "cocteleria",
                    ]:
                        # Almacenar el producto con su precio (si no está ya en el diccionario)
                        if product.name not in non_food_products:
                            non_food_products[product.name] = product.price

            # Procesar datos por día y categoría
            sales_by_day = defaultdict(
                lambda: {
                    "non_food": {
                        name: {
                            "quantity": 0,
                            "unit_price": price,
                            "total_price": 0,
                            "orders": set(),
                        }
                        for name, price in non_food_products.items()
                    },
                    "food": defaultdict(
                        lambda: {
                            "quantity": 0,
                            "unit_price": 0,
                            "total_price": 0,
                            "orders": set(),
                        }
                    ),
                }
            )

            for sale in sales:
                date_key = sale.date.strftime("%d-%m-%Y")
                for sale_product in sale.products:
                    product = sale_product.product
                    if product.category == "cocteleria":
                        continue

                    category_key = (
                        "food"
                        if product.category in ["comida", "postre", "trago"]
                        else "non_food"
                    )

                    # Agregar producto al día y categoría correspondiente
                    product_entry = sales_by_day[date_key][category_key][product.name]
                    product_entry["quantity"] += sale_product.quantity
                    if sale_product.unit_price:
                        product_entry["unit_price"] = sale_product.unit_price
                    product_entry["total_price"] += sale_product.total_price
                    product_entry["orders"].add(sale.sale_number)

            # Formatear para la plantilla
            daily_sales = []
            for date, categories in sales_by_day.items():
                non_food_list = [
                    {
                        "name": name,
                        "quantity": data["quantity"],
                        "unit_price": round(data["unit_price"], 2),
                        "total_price": round(data["total_price"], 2),
                        "orders": sorted(data["orders"]),
                    }
                    for name, data in categories["non_food"].items()
                ]
                non_food_list.sort(key=lambda x: x["name"])  # Ordenar por nombre

                food_list = [
                    {
                        "name": name,
                        "quantity": data["quantity"],
                        "unit_price": round(data["unit_price"], 2),
                        "total_price": round(data["total_price"], 2),
                        "orders": sorted(data["orders"]),
                    }
                    for name, data in categories["food"].items()
                ]
                food_list.sort(key=lambda x: x["name"])  # Ordenar por nombre

                daily_sales.append(
                    {
                        "date": date,
                        "non_food": non_food_list,
                        "food": food_list,
                    }
                )

        except ValueError:
            flash("Formato de mes inválido. Usa YYYY-MM.", "error")

    return render_template(
        "report/ipv_report.html",
        business=business,
        daily_sales=daily_sales,
        daily_sales_json=json.dumps(daily_sales),
        selected_month=selected_month,
    )


@bp.route("/export-ipv", methods=["POST"])
def export_ipv(business_id):
    business = Business.query.get_or_404(business_id)
    business_name = business.name if business.is_general else business.parent_name()
    month = request.form.get("selected_month", "_")

    try:
        # Convertir los datos JSON a una estructura de Python
        data = json.loads(request.form.get("ipv_data", "[]"))
    except json.JSONDecodeError:
        flash("Los datos recibidos no son válidos.", "error")
        return redirect(url_for("report.ipv_report", business_id=business.id))

    try:
        # Generar el reporte en MS Excel
        excel_file = generate_excel_ipv(business_name, data, month)
    except Exception as e:
        flash("Ocurrio un error al tratar de generar el reporte en MS Excel.", "error")
        return redirect(url_for("report.ipv_report", business_id=business.id))

    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"IPV_{business_name}_{month}.xlsx",
    )


@bp.route("/monthly-sales-cocteleria", methods=["GET", "POST"])
def monthly_sales_coctelera(business_id):
    # Obtener el negocio (asegúrate de que exista)
    business = Business.query.get_or_404(business_id)
    business_filters = business_service.get_parent_filters(business)

    # Inicializar variables
    daily_sales = []
    selected_month = None
    excluded_sales = []
    excluded_products = []

    if request.method == "POST":
        # Obtener el mes seleccionado desde el formulario
        selected_month = request.form.get("month")

        if not selected_month:
            flash("Por favor, selecciona un mes.", "error")
            return render_template(
                "report/monthly_sales_by_product.html",
                business=business,
                daily_sales=daily_sales,
                selected_month=selected_month,
            )

        try:

            sales, formated_daily_sales = sales_service.get_daily_sales(
                selected_month,
                business_filters["business_id"],
                business_filters.get("specific_business_id"),
            )

            # Procesar los datos
            sales_by_day = {}
            for sale in sales:
                date_key = sale.date.strftime("%Y-%m-%d")
                if date_key not in sales_by_day:
                    sales_by_day[date_key] = {
                        "total_products": 0,
                        "total_income": 0,
                        "products": {},
                    }

                # Ordenar los productos dentro de la venta por nombre
                sorted_products = sorted(
                    sale.products, key=lambda sale_detail: sale_detail.product.name
                )
                # logging.debug("Datos generados para el reporte: %s", sale.products)

                for sale_detail in sorted_products:
                    product = sale_detail.product
                    if product.category != 'cocteleria':
                        continue
                    product_key = product.name
                    if product_key not in sales_by_day[date_key]["products"]:
                        sales_by_day[date_key]["products"][product_key] = {
                            "quantity": 0,
                            "total_amount": 0,
                            "orders": set(),  # Usamos un conjunto para evitar duplicados,
                        }

                    sales_by_day[date_key]["products"][product_key][
                        "quantity"
                    ] += sale_detail.quantity
                    sales_by_day[date_key]["products"][product_key][
                        "total_amount"
                    ] += sale_detail.quantity * product.price

                    # Añadir la tupla (sale_id, sale_number) al conjunto
                    sales_by_day[date_key]["products"][product_key]["orders"].add(
                        (sale.id, sale.sale_number)
                    )

                    # Actualizar totales
                    sales_by_day[date_key]["total_products"] += sale_detail.quantity

                    sales_by_day[date_key]["total_income"] += sale_detail.quantity * product.price

            # Formatear los datos para la plantilla
            for date, data in sales_by_day.items():
                # Ordenar los productos por nombre
                sorted_products = sorted(
                    data["products"].items(),
                    key=lambda item: item[
                        0
                    ],  # Ordenar por el nombre del producto (item[0])
                )

                daily_sales.append(
                    {
                        "date": date,
                        "total_products": data["total_products"],
                        "total_income": data["total_income"],
                        "products": [
                            {
                                "name": name,
                                "quantity": product["quantity"],
                                "total_amount": product["total_amount"],
                                "orders": sorted(
                                    list(product["orders"]), key=lambda order: order[1]
                                ),
                            }
                            for name, product in sorted_products  # Usar los productos ordenados
                        ],
                    }
                )

        except ValueError:
            # Manejar errores si el formato del mes es incorrecto
            flash("Por favor, selecciona un mes válido.", "error")

    # logging.basicConfig(level=logging.DEBUG)
    # logging.debug("Datos generados para el reporte: %s", daily_sales)

    return render_template(
        "report/monthly_sales_cocteleria.html",
        business=business,
        daily_sales=daily_sales,
        daily_sales_json=json.dumps(daily_sales),
        selected_month=selected_month,
    )
