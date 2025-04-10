from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def generate_excel_sales_by_date(business, sales_data, period):
    # Crear un archivo Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumen de Ventas"

    # Definir estilos
    title_font = Font(size=16, bold=True, color="000000")
    subtitle_font = Font(size=14, bold=True, color="000000")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True, size=12, color="000000")
    total_font = Font(bold=True, size=12, color="00B050")
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    alignment_right = Alignment(horizontal="right", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )

    # Encabezado del reporte
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    sheet.cell(row=1, column=1, value=f"Reporte de Ventas - {business.name}").font = (
        title_font
    )
    sheet.cell(row=1, column=1).alignment = alignment_center

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    sheet.cell(row=2, column=1, value=f"Mes: {period}").font = subtitle_font
    sheet.cell(row=2, column=1).alignment = alignment_center

    # Espacio en blanco
    sheet.append([])
    sheet.append([])

    # Resumen general
    total_products = sum(day["total_products"] for day in sales_data)
    total_income = sum(day["total_income"] for day in sales_data)

    sheet.append(["Resumen General"])
    sheet.append(["Total de Productos Vendidos:", total_products])
    sheet.append(["Total de Ingresos Generados:", total_income])

    # Aplicar formato de moneda al total de ingresos
    sheet.cell(row=5, column=1).font = subtitle_font
    sheet.cell(row=6, column=1).font = bold_font
    sheet.cell(row=7, column=1).font = bold_font
    sheet.cell(row=6, column=2).font = bold_font
    sheet.cell(row=7, column=2).number_format = '"$ "#,##0.00'  # Formato de moneda
    sheet.cell(row=7, column=2).font = total_font

    # Espacio en blanco
    sheet.append([])
    sheet.append([])

    # Encabezados de la tabla
    headers = ["FECHA", "TOTAL PRODUCTOS", "IMPORTE TOTAL"]
    sheet.append(headers)

    # Aplicar estilos a los encabezados
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=10, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = alignment_center

    # Rellenar el archivo con los datos
    for idx, day in enumerate(sales_data, start=11):
        date = day["date"]
        total_products = day["total_products"]
        total_income = day["total_income"]

        row = [date, total_products, total_income]  # Mantener el valor numérico
        sheet.append(row)

        # Aplicar formato de moneda al importe total
        sheet.cell(row=idx, column=3).number_format = (
            '"$ "#,##0.00'  # Formato de moneda
        )
        sheet.cell(row=idx, column=3).font = bold_font

        # Aplicar bordes a las celdas
        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=idx, column=col_num)
            cell.border = border
            cell.alignment = alignment_center
            if col_num == 3:
                cell.alignment = alignment_right

    # Ajustar el ancho de las columnas automáticamente
    for col in range(1, len(headers) + 1):
        column = get_column_letter(col)
        sheet.column_dimensions[column].width = 27

    # Guardar el archivo en memoria
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    return excel_file


def generate_excel_ipv(business_name, data, month):
    """
    Genera un archivo Excel con un inventario a precio de venta (IPV) organizado por días.

    Args:
        business_name (str): Nombre del negocio.
        data (list): Lista de diccionarios que contienen los datos diarios de productos "No Comida" y "Comida".
        month (str): Mes correspondiente al reporte.

    Returns:
        BytesIO: Archivo Excel en memoria.
    """

    # Crear Excel
    workbook = Workbook()
    del workbook["Sheet"]  # Eliminar hoja por defecto

    # Estilos generales
    bold_font = Font(bold=True)
    title_font = Font(size=16, bold=True, color="000000")
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")

    header_fill = PatternFill(
        start_color="BFBFBF", end_color="BFBFBF", fill_type="solid"
    )

    # Definir estilos de borde
    thin_border_buttom = Border(bottom=Side(style="thin"))  # Borde delgado
    medium_border_buttom = Border(bottom=Side(style="medium"))  # Borde grueso

    for day in data:
        # Crear nueva hoja por día
        sheet = workbook.create_sheet(title=day["date"][:-5])

        # Configuración de la página
        sheet.page_setup.orientation = (
            sheet.ORIENTATION_LANDSCAPE
        )  # Orientación horizontal
        sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER  # Tamaño de papel A4
        # sheet.sheet_properties.pageSetUpPr.fitToPage = (
        #     True  # Ajustar al ancho de la página
        # )

        # Configuración de márgenes
        sheet.page_margins = PageMargins(
            left=0.4,  # Margen izquierdo
            right=0.4,  # Margen derecho
            top=0.4,  # Margen superior
            bottom=0.4,  # Margen inferior
            header=0.1,  # Margen del encabezado (opcional)
            footer=0.1,  # Margen del pie de página (opcional)
        )

        # Centrar la página horizontalmente
        # sheet.print_options.horizontalCentered = True

        # Repetir las primeras tres filas en cada página
        sheet.print_title_rows = "1:2"  # Las filas 1 a 3 se repetirán en cada página

        # Configurar el encabezado
        sheet.oddHeader.center.text = (
            "INVENTARIO A PRECIO DE VENTA"  # Texto a la derecha
        )
        sheet.oddHeader.right.text = (
            "pág. &[Page] de &[Pages]"  # Numeración de página a la izquierda
        )

        # Encabezados del template IPV
        # sheet.append(
        #     ["INVENTARIO A PRECIO DE VENTA", "", "", "", "", "", "", "", "", "", ""]
        # )
        # sheet.merge_cells("A1:K1")
        # sheet["A1"].font = title_font
        # sheet["A1"].alignment = center_alignment

        # Información de empresa y fecha
        sheet.append(
            [
                f"NEGOCIO: {business_name}",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                f"FECHA: {day['date']}",
                "",
            ]
        )
        sheet.merge_cells("A1:C1")
        sheet.merge_cells("J1:K1")
        sheet["A1"].font = bold_font
        sheet["A1"].alignment = left_alignment
        sheet["J1"].font = bold_font
        sheet["J1"].alignment = right_alignment

        # Ajustar la altura de la fila de encabezados (fila 3)
        sheet.row_dimensions[1].height = (
            30  # Altura en puntos (puedes ajustar este valor)
        )

        # Encabezados de columnas
        headers = [
            "Productos",
            "Al inicio",
            "Entradas",
            "Disponibles",
            "Com. Empl.",
            "Roturas",
            "A la Venta",
            "Al Final",
            "Vendido",
            "Precio",
            "Importe",
        ]
        sheet.append(headers)

        # Aplicar estilo a los encabezados
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=col_idx, value=header)
            cell.font = bold_font
            cell.border = medium_border_buttom
            cell.fill = header_fill
            if col_idx <= 9:  # Columnas desde "Al inicio" hasta "Vendido"
                cell.alignment = center_alignment
            else:  # Columnas "Precio" e "Importe"
                cell.alignment = right_alignment

        # Contador de filas para aplicar fórmulas
        row_counter = (
            3  # La fila donde comienzan los datos (después de los encabezados)
        )

        # Productos No Comida
        for product in day["non_food"]:
            sheet.append(
                [
                    product["name"],
                    " ",  # Al inicio (placeholder)
                    "",  # Entradas
                    None,  # Disponibles
                    "",  # Com. Empl.
                    "",  # Roturas
                    None,  # A la Venta
                    None,  # Al Final
                    (product["quantity"] if product["quantity"] > 0 else ""),  # Vendido
                    product["unit_price"],  # Precio
                    (
                        product["total_price"] if product["total_price"] > 0 else "-"
                    ),  # Importe
                ]
            )

            # Aplicar fórmulas en las columnas correspondientes
            formula_row = row_counter  # Fila actual

            # Fórmula para "Disponibles"
            formula_available = f'=IFERROR(IF((B{formula_row}+C{formula_row})>0,B{formula_row}+C{formula_row},""),"")'
            sheet[f"D{formula_row}"] = formula_available  # Columna "Disponibles"

            # Fórmula para "A la Venta" con SI.ERROR
            formula_for_sale = f'=IFERROR(IF((D{formula_row}-E{formula_row}-F{formula_row})>0,D{formula_row}-E{formula_row}-F{formula_row},"-"),"-")'
            sheet[f"G{formula_row}"] = formula_for_sale  # Columna "A la Venta"

            # Fórmula para "Al Final" con SI.ERROR
            formula_at_the_end = f'=IFERROR(IF((G{formula_row}-I{formula_row})>0,G{formula_row}-I{formula_row},"-"),"-")'
            sheet[f"H{formula_row}"] = formula_at_the_end  # Columna "Al Final"

            # Aplicar alineación a las celdas
            for col_idx in range(2, 10):  # Columnas desde "Al inicio" hasta "Vendido"
                sheet.cell(row=formula_row, column=col_idx).alignment = center_alignment
            for col_idx in range(10, 12):  # Columnas "Precio" e "Importe"
                sheet.cell(row=formula_row, column=col_idx).alignment = right_alignment

            # Aplicar borde inferior a la fila completa
            for col_idx in range(1, 12):  # Todas las columnas (A a K)
                sheet.cell(row=formula_row, column=col_idx).border = thin_border_buttom

            row_counter += 1  # Incrementar contador de filas

        # Aplicar borde grueso después del último producto "No Comida"
        if day["non_food"]:  # Solo si hay productos "No Comida"
            last_non_food_row = row_counter - 1
            for col_idx in range(1, 12):  # Todas las columnas (A a K)
                sheet.cell(row=last_non_food_row, column=col_idx).border = (
                    medium_border_buttom
                )

        # Productos Comida
        for product in day["food"]:
            sheet.append(
                [
                    product["name"],
                    "",  # Al inicio (placeholder)
                    "",  # Entradas
                    "",  # Disponibles
                    "",  # Com. Empl.
                    "",  # Roturas
                    "",  # A la Venta
                    "",  # Al Final
                    product["quantity"],  # Vendido
                    product["unit_price"],  # Precio
                    product["total_price"],  # Importe
                ]
            )

            # Aplicar alineación a las celdas
            for col_idx in range(2, 10):  # Columnas desde "Al inicio" hasta "Vendido"
                sheet.cell(row=row_counter, column=col_idx).alignment = center_alignment
            for col_idx in range(10, 12):  # Columnas "Precio" e "Importe"
                sheet.cell(row=row_counter, column=col_idx).alignment = right_alignment

            # Aplicar borde inferior a la fila completa
            for col_idx in range(1, 12):  # Todas las columnas (A a K)
                sheet.cell(row=row_counter, column=col_idx).border = thin_border_buttom

            row_counter += 1  # Incrementar contador de filas

        # Aplicar borde grueso después del último producto "Comida"
        if day["food"]:  # Solo si hay productos "Comida"
            last_food_row = row_counter - 1
            for col_idx in range(1, 12):  # Todas las columnas (A a K)
                sheet.cell(row=last_food_row, column=col_idx).border = (
                    medium_border_buttom
                )

        # Ajustar ancho de las columnas
        column_widths = [30, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10]
        for idx, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[chr(64 + idx)].width = width

    # Guardar y enviar
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    return excel_file
