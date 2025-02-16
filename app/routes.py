from sqlite3 import IntegrityError
from flask import (
    Blueprint,
    json,
    render_template,
    request,
    redirect,
    send_file,
    url_for,
    flash,
    current_app,
    jsonify,
    session,
)
from werkzeug.utils import secure_filename
from app.forms import BusinessForm, EditSaleProductForm
from app.models import db, Business, Product, Sale, SaleProduct
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from collections import defaultdict
import matplotlib
from io import BytesIO
from openpyxl import Workbook
import json
import logging


matplotlib.use("Agg")
import matplotlib.pyplot as plt
import os

bp = Blueprint("main", __name__)
logging.basicConfig(level=logging.DEBUG)


@bp.route("/")
def index():
    # Esta función muestra la lista de negocios en la página principal
    business_list = Business.query.all()
    # Realiza una consulta a todos los registros de negocios en la base de datos
    return render_template("index.html", business_list=business_list)


@bp.route("/business_list", methods=["GET", "POST"])
def business_list():
    # Inicializar el formulario para la entrada de datos
    form = BusinessForm()

    if form.validate_on_submit():
        # Procesar los datos del formulario si hay validación exitosa
        name = form.name.data
        description = form.description.data
        logo_file = form.logo.data

        # Manejar el archivo de logo si se proporciona
        logo_path = None
        if logo_file and logo_file.filename != "":
            # Obtener el nombre seguro del archivo y construir la ruta
            filename = secure_filename(logo_file.filename)
            logo_path = os.path.join("images", "logos", filename)

            # Construir la ruta completa para guardar el archivo
            full_path = os.path.join(
                current_app.root_path, "static", "images", "logos", filename
            )
            logo_file.save(full_path)

        # Crear un nuevo negocio con los datos proporcionados
        new_business = Business(name=name, description=description, logo=logo_path)

        db.session.add(new_business)
        db.session.commit()

        flash("Negocio agregado correctamente", "success")
        return redirect(url_for("main.business_list"))

    # Si no hay validación o es una solicitud GET, mostrar la lista de negocios
    business_list = Business.query.all()
    return render_template("business.html", business_list=business_list, form=form)


@bp.route("/business/<int:business_id>", methods=["GET", "POST"])
def business_detail_or_edit(business_id):
    # Obtener el negocio por su ID
    business = Business.query.get_or_404(business_id)

    # Determinar si estamos en modo "editar"
    edit = request.args.get("edit", False)  # Por defecto, edit es False

    if edit:
        # Crear una instancia del formulario con los datos actuales del negocio
        form = BusinessForm(obj=business)

        if form.validate_on_submit():
            # Actualizar los datos del negocio
            business.name = form.name.data
            business.description = form.description.data

            # Manejar el logo si se proporciona uno nuevo
            logo_file = form.logo.data
            if logo_file and logo_file.filename != "":
                filename = secure_filename(logo_file.filename)
                logo_path = os.path.join("images", "logos", filename)
                full_path = os.path.join(
                    current_app.root_path, "static", "images", "logos", filename
                )
                logo_file.save(full_path)
                business.logo = logo_path

            form.populate_obj(business)
            # Guardar los cambios en la base de datos
            db.session.commit()

            flash("Negocio actualizado correctamente", "success")
            return redirect(
                url_for("main.business_detail_or_edit", business_id=business.id)
            )

        return render_template(
            "business_detail_or_edit.html", business=business, form=form, edit=True
        )

    # Si no estamos en modo "editar", mostrar los detalles
    return render_template(
        "business_detail_or_edit.html", business=business, edit=False
    )


@bp.route("/business/<int:business_id>/dashboard")
def business_dashboard(business_id):
    monthly_totals = {}
    total_general = 0
    business = Business.query.get_or_404(business_id)

    sales = (
        Sale.query.filter_by(business_id=business.id).order_by(Sale.date.desc()).all()
    )

    for sale in sales:
        month = sale.date.strftime("%Y-%m")
        total = sum(
            sale_product.product.price * sale_product.quantity
            for sale_product in sale.products
        )
        monthly_totals[month] = monthly_totals.get(month, 0) + total
        total_general += total
    months = list(monthly_totals.keys())
    totals = list(monthly_totals.values())

    # Generar el gráfico
    plt.figure(figsize=(8, 4))
    plt.bar(months, totals, color="skyblue")
    plt.title("Desempeño Mensual")
    plt.xlabel("Mes")
    plt.ylabel("Total Acumulado")

    # Guardar el gráfico
    chart_filename = f"chart_{business_id}.png"
    chart_path = os.path.join("images", "charts", chart_filename)
    full_chart_path = os.path.join(
        current_app.root_path, "static", "images", "charts", chart_filename
    )
    plt.savefig(full_chart_path)
    plt.close()

    return render_template(
        "business_dashboard.html",
        business=business,
        monthly_totals=monthly_totals,
        total_general=total_general,
        chart_path=chart_path,
    )


@bp.route("/business/<int:business_id>/products", methods=["GET", "POST"])
def products(business_id):
    business = Business.query.get_or_404(business_id)

    if request.method == "POST":
        name = request.form["name"]
        price = float(request.form["price"])
        new_product = Product(name=name, price=price, business_id=business.id)
        db.session.add(new_product)
        db.session.commit()
        flash("Producto agregado correctamente", "success")
        return redirect(url_for("main.products", business_id=business.id))

    products_list = (
        Product.query.filter_by(business_id=business.id).order_by(Product.name).all()
    )
    return render_template("products.html", business=business, products=products_list)


@bp.route("/business/<int:business_id>/products/<int:product_id>", methods=["POST"])
def update_product(business_id, product_id):
    business = Business.query.get_or_404(business_id)

    product = Product.query.get_or_404(product_id)
    product.name = request.form["name"]
    product.price = float(request.form["price"])
    db.session.commit()
    flash("Producto actualizado correctamente", "success")
    return redirect(url_for("main.products", business_id=business.id))


@bp.route("/business/<int:business_id>/sales", methods=["GET", "POST"])
def sales(business_id):
    business = Business.query.get_or_404(business_id)

    if request.method == "POST":
        date_str = request.form["date"]
        date = datetime.strptime(date_str, "%Y-%m-%d").date()

        # Obtener el último número de venta para esta ubicación y año
        last_sale = (
            Sale.query.filter_by(
                business_id=business.id,
                # date=date  # Asegúrate de que sea del mismo año si es necesario
            )
            .order_by(Sale.sale_number.desc())
            .first()
        )

        year = date.year

        if last_sale and last_sale.year == year:
            new_sale_number = last_sale.sale_number + 1
        else:
            new_sale_number = 1

        # Verificar que el par (business_id, sale_number) no exista
        try:
            new_sale = Sale(
                sale_number=new_sale_number,
                date=date,
                year=date.year,
                business_id=business.id,
            )
            db.session.add(new_sale)
            db.session.commit()
            flash("Venta creada correctamente", "success")
            return redirect(
                url_for(
                    "main.add_sale_products",
                    business_id=business.id,
                    sale_id=new_sale.id,
                )
            )
        except IntegrityError:
            db.session.rollback()
            flash("Error al crear la venta. El número ya existe.", "danger")
            return redirect(url_for("main.sales", business_id=business.id))

    # Obtener todas las órdenes
    all_sales = (
        Sale.query.filter_by(business_id=business.id).order_by(Sale.date.desc()).all()
    )

    # Agrupar órdenes por mes y fecha
    sales_by_months = defaultdict(
        lambda: defaultdict(
            lambda: {"total_products": 0, "total_income": 0, "sales": []}
        )
    )
    month_totals = defaultdict(float)

    for sale in all_sales:
        month_key = sale.date.strftime("%Y-%m")
        date_key = sale.date.strftime("%Y-%m-%d")

        # Calcular totales por venta
        total_products = sum(sale_product.quantity for sale_product in sale.products)
        total_income = sum(
            sale_product.product.price * sale_product.quantity
            for sale_product in sale.products
        )

        # Agregar datos a la estructura
        sales_by_months[month_key][date_key]["total_products"] += total_products
        sales_by_months[month_key][date_key]["total_income"] += total_income
        sales_by_months[month_key][date_key]["sales"].append(sale)

        # Acumular totales por mes
        month_totals[month_key] += total_income

    # Convertir defaultdict a diccionarios regulares para Jinja2
    sales_by_months = {month: dict(dates) for month, dates in sales_by_months.items()}
    month_totals = dict(month_totals)

    return render_template(
        "sales.html",
        business=business,
        sales_by_months=sales_by_months,
        month_totals=month_totals,
    )


@bp.route(
    "/business/<int:business_id>/sales/<int:sale_id>/add_products",
    methods=["GET", "POST"],
)
def add_sale_products(business_id, sale_id):
    business = Business.query.get_or_404(business_id)

    sale = Sale.query.get_or_404(sale_id)
    if request.method == "POST":
        product_id = int(request.form["product_id"])
        quantity = int(request.form["quantity"])
        sale_product = SaleProduct(
            sale_id=sale.id, product_id=product_id, quantity=quantity
        )
        db.session.add(sale_product)
        db.session.commit()
        flash("Producto agregado a la venta", "success")
        return redirect(
            url_for("main.add_sale_products", business_id=business.id, sale_id=sale.id)
        )

    products_list = (
        Product.query.filter_by(business_id=business.id)
        .order_by(Product.name.asc())
        .all()
    )

    total = sum(
        sale_product.product.price * sale_product.quantity
        for sale_product in sale.products
    )

    return render_template(
        "add_sale.html",
        business=business,
        sale=sale,
        products=products_list,
        total=total,
    )


@bp.route(
    "/business/<int:business_id>/sale/<int:sale_id>/edit-product/<int:product_id>",
    methods=["GET", "POST"],
)
def edit_sale_product(business_id, sale_id, product_id):
    # Obtener el negocio
    business = Business.query.get_or_404(business_id)

    # Obtener la venta y el producto específico
    sale = Sale.query.get_or_404(sale_id)
    sale_product = SaleProduct.query.filter_by(
        sale_id=sale_id, product_id=product_id
    ).first_or_404()

    # Crear el formulario
    form = EditSaleProductForm(
        obj=sale_product
    )  # Prellenar el formulario con los datos actuales

    if form.validate_on_submit():
        # Actualizar los datos del producto en la venta
        sale_product.quantity = form.quantity.data
        db.session.commit()

        flash("El producto ha sido actualizado correctamente.", "success")
        return redirect(
            url_for("main.sale_details", business_id=business_id, sale_id=sale_id)
        )

    return render_template(
        "edit_sale_product.html",
        business=business,
        sale=sale,
        sale_product=sale_product,
        form=form,
    )


@bp.route("/business/<int:business_id>/sales/<int:sale_id>")
def sale_details(business_id, sale_id):
    business = Business.query.get_or_404(business_id)

    sale = Sale.query.get_or_404(sale_id)
    total = sum(
        sale_product.product.price * sale_product.quantity
        for sale_product in sale.products
    )
    return render_template(
        "sale_details.html", business=business, sale=sale, total=total
    )


@bp.route(
    "/business/<int:business_id>/sale/<int:sale_id>/remove-product/<int:product_id>",
    methods=["POST"],
)
def remove_product_from_sale(business_id, sale_id, product_id):
    # Obtener el negocio
    business = Business.query.get_or_404(business_id)

    # Obtener la venta y el producto específico
    sale_product = SaleProduct.query.filter_by(
        sale_id=sale_id, product_id=product_id
    ).first_or_404()

    # Acceder al nombre del producto antes de eliminarlo
    product_name = sale_product.product.name

    # Eliminar el producto de la venta
    db.session.delete(sale_product)
    db.session.commit()

    flash(
        f"El producto '{sale_product.product.name}' ha sido eliminado de la venta.",
        "success",
    )
    return redirect(
        url_for("main.sale_details", business_id=business_id, sale_id=sale_id)
    )


@bp.route("/business/<int:business_id>/monthly-report", methods=["GET", "POST"])
def monthly_report(business_id):
    # Obtener el negocio (asegúrate de que exista)
    business = Business.query.get_or_404(business_id)

    # Inicializar variables
    daily_sales = []
    selected_month = None
    excluded_sales = []
    excluded_products = []

    if request.method == "POST":
        # Obtener el mes seleccionado desde el formulario
        month_str = request.form.get("month")

        if not month_str:
            flash("Por favor, selecciona un mes.", "error")
            return render_template(
                "monthly_report.html",
                business=business,
                daily_sales=daily_sales,
                selected_month=selected_month,
            )

        try:
            # Convertir el mes a un objeto datetime
            selected_month = datetime.strptime(month_str, "%Y-%m").date()
            start_date = selected_month.replace(day=1)
            end_date = start_date + relativedelta(months=1, days=-1)

            # Consultar las órdenes dentro del rango de fechas
            sales = (
                Sale.query.filter(
                    Sale.business_id == business.id,
                    Sale.date.between(start_date, end_date),
                )
                .join(Sale.products)  # Join con SaleProduct
                .join(SaleProduct.product)  # Join con Product
                .order_by(
                    Sale.date.asc(),  # Primero por fecha
                )
                .all()
            )

            # Procesar los datos
            sales_by_day = {}
            for sale in sales:
                date_key = sale.date.strftime("%Y-%m-%d")
                if date_key not in sales_by_day:
                    sales_by_day[date_key] = {
                        "total_products": 0,
                        "total_income": 0,
                        "products": {},
                    }

                # Ordenar los productos dentro de la venta por nombre
                sorted_products = sorted(
                    sale.products, key=lambda sale_product: sale_product.product.name
                )
                # logging.debug("Datos generados para el reporte: %s", sale.products)

                for sale_product in sorted_products:
                    product_key = sale_product.product.name
                    if product_key not in sales_by_day[date_key]["products"]:
                        sales_by_day[date_key]["products"][product_key] = {
                            "quantity": 0,
                            "total_amount": 0,
                            "orders": set(),  # Usamos un conjunto para evitar duplicados,
                        }

                    sales_by_day[date_key]["products"][product_key][
                        "quantity"
                    ] += sale_product.quantity
                    sales_by_day[date_key]["products"][product_key]["total_amount"] += (
                        sale_product.quantity * sale_product.product.price
                    )

                    # Añadir la tupla (sale_id, sale_number) al conjunto
                    sales_by_day[date_key]["products"][product_key]["orders"].add(
                        (sale.id, sale.sale_number)
                    )

                # Actualizar totales
                sales_by_day[date_key]["total_products"] += sum(
                    sale_product.quantity for sale_product in sale.products
                )
                sales_by_day[date_key]["total_income"] += sum(
                    sale_product.quantity * sale_product.product.price
                    for sale_product in sale.products
                )

            # Formatear los datos para la plantilla
            for date, data in sales_by_day.items():
                # Ordenar los productos por nombre
                sorted_products = sorted(
                    data["products"].items(),
                    key=lambda item: item[
                        0
                    ],  # Ordenar por el nombre del producto (item[0])
                )

                daily_sales.append(
                    {
                        "date": date,
                        "total_products": data["total_products"],
                        "total_income": data["total_income"],
                        "products": [
                            {
                                "name": name,
                                "quantity": product["quantity"],
                                "total_amount": product["total_amount"],
                                "orders": sorted(
                                    list(product["orders"]), key=lambda order: order[1]
                                ),
                            }
                            for name, product in sorted_products  # Usar los productos ordenados
                        ],
                    }
                )

        except ValueError:
            # Manejar errores si el formato del mes es incorrecto
            flash("Por favor, selecciona un mes válido.", "error")

    # logging.basicConfig(level=logging.DEBUG)
    # logging.debug("Datos generados para el reporte: %s", daily_sales)

    return render_template(
        "monthly_report.html",
        business=business,
        daily_sales=daily_sales,
        daily_sales_json=json.dumps(daily_sales),
        selected_month=selected_month,
    )


@bp.route("/business/<int:business_id>/monthly-report-by-date", methods=["GET", "POST"])
def monthly_report_by_date(business_id):
    # Obtener el negocio
    business = Business.query.get_or_404(business_id)

    # Inicializar variables
    all_sales = []
    daily_sales = []
    excluded_sales = [  # Ventas excluidas almacenadas en la sesión
        int(excluded_sale) for excluded_sale in session.get("excluded_sales", [])
    ]
    selected_month = None

    if request.method == "POST":
        # Obtener el mes seleccionado desde el formulario
        month_str = request.form.get("month")
        if not month_str:
            flash("Por favor, selecciona un mes.", "error")
            return render_template(
                "monthly_report_by_date.html",
                business=business,
                daily_sales=daily_sales,
                excluded_sales=excluded_sales,
                selected_month=selected_month,
            )

        try:
            # Convertir el mes a un objeto datetime
            selected_month = datetime.strptime(month_str, "%Y-%m").date()
            start_date = selected_month.replace(day=1)
            end_date = start_date + relativedelta(months=1, days=-1)

            # Consultar todas las órdenes dentro del rango de fechas
            all_sales = (
                Sale.query.filter(
                    Sale.business_id == business.id,
                    Sale.date.between(start_date, end_date),
                )
                .join(Sale.products)  # Join con SaleProduct
                .join(SaleProduct.product)  # Join con Product
                .order_by(Sale.date.asc())  # Ordenar por fecha
                .all()
            )

            # Filtrar las ventas excluidas
            filtered_sales = [
                sale for sale in all_sales if sale.sale_number not in excluded_sales
            ]

            # Procesar los datos
            sales_by_day = defaultdict(
                lambda: {"sales": [], "total_products": 0, "total_income": 0}
            )

            for sale in filtered_sales:
                date_key = sale.date.strftime("%Y-%m-%d")

                # Calcular totales para esta venta
                total_products = sum(sp.quantity for sp in sale.products)
                total_income = sum(
                    sp.quantity * sp.product.price for sp in sale.products
                )
                sorted_products = sorted(
                    sale.products,
                    key=lambda sale_product: sale_product.product.name,
                )
                # Crear un diccionario para acumular las cantidades por producto
                reduced_products_dict = defaultdict(
                    lambda: {"quantity": 0, "import": 0}
                )

                # Recorrer la lista de productos y acumular las cantidades
                for sale_product in sorted_products:
                    name = sale_product.product.name
                    quantity = sale_product.quantity
                    price = sale_product.product.price

                    # Acumular la cantidad y asignar el precio (asumiendo que el precio es el mismo para cada producto)
                    reduced_products_dict[name]["quantity"] += quantity
                    reduced_products_dict[name]["import"] = quantity * price

                # Convertir el diccionario a una lista de productos reducidos
                saled_products = [
                    {
                        "name": name,
                        "quantity": data["quantity"],
                        "import": data["import"],
                    }
                    for name, data in reduced_products_dict.items()
                ]

                # Agregar la venta a la lista de ventas del día
                sales_by_day[date_key]["sales"].append(
                    {
                        "sale_number": sale.sale_number,
                        "total_products": total_products,
                        "total_income": total_income,
                        "products": saled_products,
                    }
                )

                # Actualizar totales acumulados para el día
                sales_by_day[date_key]["total_products"] += total_products
                sales_by_day[date_key]["total_income"] += total_income

            # Formatear los datos para la plantilla
            for date, data in sales_by_day.items():
                daily_sales.append(
                    {
                        "date": date,
                        "sales": data["sales"],
                        "total_products": data["total_products"],
                        "total_income": data["total_income"],
                    }
                )

        except ValueError:
            # Manejar errores si el formato del mes es incorrecto
            flash("Por favor, selecciona un mes válido.", "error")

    # Logging para depuración
    # logging.debug("Ventas en json: %s", json.dumps(daily_sales))

    return render_template(
        "monthly_report_by_date.html",
        business=business,
        all_sales=all_sales,
        daily_sales=daily_sales,
        daily_sales_json=json.dumps(daily_sales),
        excluded_sales=excluded_sales,
        selected_month=selected_month,
    )


@bp.route("/exclude-sales", methods=["POST"])
def exclude_sales():
    data = request.get_json()
    selected_sales = data.get("sales", [])

    # Guardar las ventas excluidas en la sesión
    session["excluded_sales"] = selected_sales

    # Logging para depuración
    # logging.debug("Ventas excluidas en exclude_sales: %s", session["excluded_sales"])

    return jsonify({"message": "Ventas excluidas correctamente"}), 200


@bp.route("/business/<int:business_id>/export-to-excel-sales-by-day", methods=["POST"])
def export_to_excel_sales_by_day(business_id):
    # Obtener el negocio
    business = Business.query.get_or_404(business_id)

    # Recuperar los datos del reporte mensual
    selected_month = request.form.get("selected_month")
    daily_sales_json = request.form.get("daily_sales_export")
    if not daily_sales_json:
        flash("No hay datos disponibles para exportar.", "error")
        return redirect(url_for("main.monthly_report", business_id=business_id))

    try:
        # Convertir los datos JSON a una estructura de Python
        data = json.loads(daily_sales_json)
    except json.JSONDecodeError:
        flash("Los datos recibidos no son válidos.", "error")
        return redirect(url_for("main.monthly_report", business_id=business_id))

    # Crear un archivo Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Mensual"

    # Encabezados
    headers = [
        "Fecha",
        "Orden",
        "Producto",
        "Cantidad",
        "Importe",
    ]
    sheet.append(headers)

    # Rellenar el archivo con los datos
    for day in data:
        date = day["date"]
        for sale in day["sales"]:
            sale_number = sale["sale_number"]
            for product in sale["products"]:
                row = [
                    date,
                    sale_number,
                    product["name"],
                    product["quantity"],
                    product["import"],
                ]
                sheet.append(row)

    # Guardar el archivo en memoria
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    # Enviar el archivo como respuesta
    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{selected_month}_reporte_mensual_por_dia_{business.name}.xlsx",
    )


@bp.route(
    "/business/<int:business_id>/export-to-excel-sales-by-product", methods=["POST"]
)
def export_to_excel_sales_by_product(business_id):
    # Obtener el negocio
    business = Business.query.get_or_404(business_id)

    # Recuperar los datos del reporte mensual
    daily_sales_json = request.form.get("daily_sales_export")
    if not daily_sales_json:
        flash("No hay datos disponibles para exportar.", "error")
        return redirect(url_for("main.monthly_report", business_id=business_id))

    try:
        # Convertir los datos JSON a una estructura de Python
        data = json.loads(daily_sales_json)
    except json.JSONDecodeError:
        flash("Los datos recibidos no son válidos.", "error")
        return redirect(url_for("main.monthly_report", business_id=business_id))

    # Procesar los datos para agrupar por producto
    product_summary = defaultdict(
        lambda: {"quantity": 0, "total_amount": 0, "orders": set()}
    )

    for day in data:
        if "products" not in day:
            flash("Los datos recibidos no tienen el formato esperado.", "error")
            return redirect(url_for("main.monthly_report", business_id=business_id))

        for product in day["products"]:
            if (
                "name" not in product
                or "quantity" not in product
                or "total_amount" not in product
                or "orders" not in product
            ):
                flash("Los datos recibidos no tienen el formato esperado.", "error")
                return redirect(url_for("main.monthly_report", business_id=business_id))

            product_name = product["name"]
            product_summary[product_name]["quantity"] += product["quantity"]
            product_summary[product_name]["total_amount"] += product["total_amount"]

            # Agregar los números de orden al conjunto
            for order in product["orders"]:
                product_summary[product_name]["orders"].add(order[1])

    # Logging de los datos generados
    # logging.debug("Datos generados para el reporte: %s", dict(product_summary))

    # Crear un archivo Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Mensual"

    # Encabezados
    headers = ["PRODUCTO", "CANTIDAD", "IMPORTE", "ORDENES"]
    sheet.append(headers)

    # Rellenar el archivo con los datos
    for product_name, details in product_summary.items():
        row = [
            product_name,
            details["quantity"],
            details["total_amount"],
            ", ".join(
                f"[{order}]" for order in sorted(list(details["orders"]))
            ),  # Formato [1], [2], [3]
        ]
        sheet.append(row)

    # Guardar el archivo en memoria
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    # Enviar el archivo como respuesta
    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"reporte_mensual_por_producto_{business.name}.xlsx",
    )


@bp.route("/business/<int:business_id>/export-to-excel-exclude-sales", methods=["POST"])
def export_to_excel_exclude_sales(business_id):
    # Obtener el negocio
    business = Business.query.get_or_404(business_id)

    # Recuperar las ventas excluidas de la sesión
    excluded_sales = session.get("excluded_sales", [])

    # Obtener todas las ventas del negocio
    all_sales = Sale.query.filter_by(business_id=business.id).all()

    # Filtrar las ventas excluidas
    filtered_sales = [sale for sale in all_sales if sale.id not in excluded_sales]

    # Crear un archivo Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = datetime.now().strftime("%m-%Y")

    # Encabezados
    headers = [
        "FECHA",
        "CANTIDAD DE PRODUCTOS POR FECHA",
        "IMPORTE POR FECHA",
        "VENTAS POR FECHA",
    ]
    sheet.append(headers)

    # Rellenar el archivo con los datos
    for sale in filtered_sales:
        row = [
            sale.date,
            sum(sp.quantity for sp in sale.products),
            sum(sp.quantity * sp.product.price for sp in sale.products),
            len(sale.products),
        ]
        sheet.append(row)

    # Guardar el archivo en memoria
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    # Enviar el archivo como respuesta
    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"reporte-{datetime.now().strftime('%m-%Y')}.xlsx",
    )
