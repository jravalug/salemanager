from datetime import datetime
from io import BytesIO
import json
from typing import List
from collections import defaultdict
from sqlalchemy import func
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook

from app.models import (
    Sale,
    Business,
    IncomeEvent,
    FinancialLedgerEntry,
    FiscalIncomeEntry,
)
from app.repositories.income_repository import IncomeRepository
from app.services.business_service import BusinessService
from app.utils.slug_utils import get_business_by_slugs
from app.utils import (
    format_daily_sales,
)


class IncomeReportService:
    def __init__(self):
        """Inicializa dependencias de reportes de ventas."""
        self.repository = IncomeRepository()
        self.business_service = BusinessService()

    def resolve_business_scope(self, client_slug: str, business_slug: str):
        """Resuelve negocio y filtros de alcance a partir de slugs."""
        business = get_business_by_slugs(client_slug, business_slug)
        if not business:
            raise ValueError("Negocio no encontrado")

        business_filters = self.business_service.get_parent_filters(business)
        return business, business_filters

    @staticmethod
    def parse_json_payload(raw_payload: str | None, require_non_empty: bool = False):
        """Parsea un payload JSON y valida opcionalmente que no esté vacío."""
        try:
            data = json.loads(raw_payload or "[]")
        except json.JSONDecodeError:
            raise ValueError("Los datos recibidos no son válidos.")

        if require_non_empty and not data:
            raise ValueError("No hay datos disponibles para exportar.")

        return data

    @staticmethod
    def _month_date_range(month_str: str):
        """Convierte `YYYY-MM` a rango de fechas del mes seleccionado."""
        try:
            selected_month = datetime.strptime(month_str, "%Y-%m").date()
        except ValueError:
            raise ValueError("El mes debe estar en formato YYYY-MM.")

        start_date = selected_month.replace(day=1)
        end_date = start_date + relativedelta(months=1, days=-1)
        return selected_month, start_date, end_date

    def _get_sales_for_month_by_scope(
        self,
        month_str: str,
        business_filters: dict,
    ) -> list[Sale]:
        """Obtiene ventas mensuales del repositorio según alcance de negocio."""
        _, start_date, end_date = self._month_date_range(month_str)
        return self.repository.get_sales_for_month(
            business_id=business_filters["business_id"],
            specific_business_id=business_filters.get("specific_business_id"),
            start_date=start_date,
            end_date=end_date,
        )

    @staticmethod
    def _parse_date(date_str: str, field_name: str):
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            raise ValueError(f"{field_name} debe estar en formato YYYY-MM-DD.")

    def _parse_date_range(self, start_date: str | None, end_date: str | None):
        if bool(start_date) != bool(end_date):
            raise ValueError("Debes enviar start_date y end_date juntos.")

        if not start_date and not end_date:
            return None, None

        parsed_start = self._parse_date(start_date, "start_date")
        parsed_end = self._parse_date(end_date, "end_date")
        if parsed_start > parsed_end:
            raise ValueError("start_date no puede ser mayor que end_date.")

        return parsed_start, parsed_end

    @staticmethod
    def _validate_regime_filter(regime: str | None):
        if not regime:
            return None

        normalized = regime.strip().lower()
        allowed = {
            FinancialLedgerEntry.REGIME_FISCAL,
            FinancialLedgerEntry.REGIME_FINANCIAL,
        }
        if normalized not in allowed:
            raise ValueError("regime debe ser 'fiscal' o 'financiera'.")
        return normalized

    @staticmethod
    def _resolve_aging_bucket(days_pending: int) -> str:
        if days_pending <= 30:
            return "0-30"
        if days_pending <= 60:
            return "31-60"
        if days_pending <= 90:
            return "61-90"
        return "90+"

    def get_financial_ledger_report(
        self,
        business_id: int,
        start_date: str | None = None,
        end_date: str | None = None,
        regime: str | None = None,
    ):
        parsed_start, parsed_end = self._parse_date_range(start_date, end_date)
        regime_filter = self._validate_regime_filter(regime)

        query = FinancialLedgerEntry.query.filter_by(business_id=business_id)
        if parsed_start and parsed_end:
            query = query.filter(
                FinancialLedgerEntry.recognition_date.between(parsed_start, parsed_end)
            )
        if regime_filter:
            query = query.filter(FinancialLedgerEntry.regime == regime_filter)

        entries = query.order_by(
            FinancialLedgerEntry.recognition_date.asc(),
            FinancialLedgerEntry.id.asc(),
        ).all()

        return {
            "entries": [
                {
                    "id": entry.id,
                    "income_event_id": entry.income_event_id,
                    "recognition_date": entry.recognition_date.strftime("%Y-%m-%d"),
                    "amount": float(entry.amount or 0),
                    "regime": entry.regime,
                    "source_ref": entry.source_ref,
                }
                for entry in entries
            ],
            "totals": {
                "count": len(entries),
                "amount": round(sum(float(entry.amount or 0) for entry in entries), 2),
            },
        }

    def get_fiscal_ledger_report(
        self,
        business_id: int,
        start_date: str | None = None,
        end_date: str | None = None,
        regime: str | None = None,
    ):
        parsed_start, parsed_end = self._parse_date_range(start_date, end_date)
        regime_filter = self._validate_regime_filter(regime)

        query = FiscalIncomeEntry.query.filter_by(business_id=business_id)
        if parsed_start and parsed_end:
            query = query.filter(
                FiscalIncomeEntry.recognition_date.between(parsed_start, parsed_end)
            )
        if regime_filter:
            query = query.filter(FiscalIncomeEntry.regime == regime_filter)

        entries = query.order_by(
            FiscalIncomeEntry.recognition_date.asc(),
            FiscalIncomeEntry.id.asc(),
        ).all()

        return {
            "entries": [
                {
                    "id": entry.id,
                    "income_event_id": entry.income_event_id,
                    "recognition_date": entry.recognition_date.strftime("%Y-%m-%d"),
                    "amount": float(entry.amount or 0),
                    "regime": entry.regime,
                    "source_ref": entry.source_ref,
                }
                for entry in entries
            ],
            "totals": {
                "count": len(entries),
                "amount": round(sum(float(entry.amount or 0) for entry in entries), 2),
            },
        }

    def get_pending_aging_report(
        self,
        business_id: int,
        as_of_date: str | None = None,
        regime: str | None = None,
    ):
        business = Business.query.get(business_id)
        if not business:
            raise ValueError("Negocio no encontrado")

        regime_filter = self._validate_regime_filter(regime)
        business_regime = (business.client.accounting_regime or "").strip().lower()
        if regime_filter and regime_filter != business_regime:
            return {
                "as_of_date": datetime.today().strftime("%Y-%m-%d"),
                "entries": [],
                "aging": {"0-30": 0.0, "31-60": 0.0, "61-90": 0.0, "90+": 0.0},
                "totals": {"count": 0, "amount": 0.0},
            }

        effective_date = (
            self._parse_date(as_of_date, "as_of_date")
            if as_of_date
            else datetime.today().date()
        )

        pending_events = (
            IncomeEvent.query.filter_by(
                business_id=business_id,
                collection_status=IncomeEvent.STATUS_PENDING,
            )
            .order_by(IncomeEvent.event_date.asc(), IncomeEvent.id.asc())
            .all()
        )

        entries = []
        aging_totals = {"0-30": 0.0, "31-60": 0.0, "61-90": 0.0, "90+": 0.0}
        for event in pending_events:
            days_pending = max((effective_date - event.event_date).days, 0)
            bucket = self._resolve_aging_bucket(days_pending)
            amount = float(event.amount or 0)
            aging_totals[bucket] += amount
            entries.append(
                {
                    "income_event_id": event.id,
                    "event_date": event.event_date.strftime("%Y-%m-%d"),
                    "days_pending": days_pending,
                    "aging_bucket": bucket,
                    "amount": amount,
                    "payment_channel": event.payment_channel,
                    "origin_type": event.origin_type,
                    "source_ref": event.source_ref,
                }
            )

        return {
            "as_of_date": effective_date.strftime("%Y-%m-%d"),
            "entries": entries,
            "aging": {key: round(value, 2) for key, value in aging_totals.items()},
            "totals": {
                "count": len(entries),
                "amount": round(sum(item["amount"] for item in entries), 2),
            },
        }

    def get_regime_compliance_report(
        self,
        business_id: int,
        start_date: str | None = None,
        end_date: str | None = None,
        regime: str | None = None,
    ):
        parsed_start, parsed_end = self._parse_date_range(start_date, end_date)
        business = Business.query.get(business_id)
        if not business:
            raise ValueError("Negocio no encontrado")

        applicable_regime = (business.client.accounting_regime or "").strip().lower()
        requested_regime = self._validate_regime_filter(regime)
        if requested_regime and requested_regime != applicable_regime:
            raise ValueError(
                "El régimen solicitado no coincide con el régimen contable del cliente."
            )

        financial_query = FinancialLedgerEntry.query.filter_by(business_id=business_id)
        fiscal_query = FiscalIncomeEntry.query.filter_by(business_id=business_id)
        pending_query = IncomeEvent.query.filter_by(
            business_id=business_id,
            collection_status=IncomeEvent.STATUS_PENDING,
        )

        if parsed_start and parsed_end:
            financial_query = financial_query.filter(
                FinancialLedgerEntry.recognition_date.between(parsed_start, parsed_end)
            )
            fiscal_query = fiscal_query.filter(
                FiscalIncomeEntry.recognition_date.between(parsed_start, parsed_end)
            )
            pending_query = pending_query.filter(
                IncomeEvent.event_date.between(parsed_start, parsed_end)
            )

        financial_total = (
            financial_query.with_entities(
                func.coalesce(func.sum(FinancialLedgerEntry.amount), 0.0)
            ).scalar()
            or 0.0
        )
        fiscal_total = (
            fiscal_query.with_entities(
                func.coalesce(func.sum(FiscalIncomeEntry.amount), 0.0)
            ).scalar()
            or 0.0
        )
        pending_total = (
            pending_query.with_entities(
                func.coalesce(func.sum(IncomeEvent.amount), 0.0)
            ).scalar()
            or 0.0
        )

        applicable_total = (
            float(fiscal_total)
            if applicable_regime == FiscalIncomeEntry.REGIME_FISCAL
            else float(financial_total)
        )

        return {
            "business_id": business_id,
            "client_id": business.client.id,
            "applicable_regime": applicable_regime,
            "period": {
                "start_date": (
                    parsed_start.strftime("%Y-%m-%d") if parsed_start else None
                ),
                "end_date": parsed_end.strftime("%Y-%m-%d") if parsed_end else None,
            },
            "totals": {
                "financial_amount": round(float(financial_total), 2),
                "fiscal_amount": round(float(fiscal_total), 2),
                "pending_amount": round(float(pending_total), 2),
                "applicable_book_amount": round(float(applicable_total), 2),
            },
            "compliance": {
                "is_regime_aligned": True,
                "message": "Reporte generado usando el régimen contable aplicable del cliente.",
            },
        }

    @staticmethod
    def generate_excel_tabular_report(
        title: str,
        headers: list[str],
        rows: list[list],
    ):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title[:31] if title else "Reporte"

        sheet.append(headers)
        for row in rows:
            sheet.append(row)

        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        return excel_file

    def get_daily_sales(
        self,
        month_str: str,
        business_id: int,
        specific_business_id: int | None = None,
    ):
        """Obtiene las ventas diarias agrupadas y procesadas."""
        _, start_date, end_date = self._month_date_range(month_str)

        # Obtener las ventas del repositorio
        all_daily_sales = self.repository.get_sales_for_month(
            business_id, specific_business_id, start_date, end_date
        )

        # Procesar las ventas
        grouped_sales_by_day = self._group_sales_by_day(all_daily_sales)
        formatted_daily_sales = format_daily_sales(grouped_sales_by_day)
        return all_daily_sales, formatted_daily_sales

    def get_monthly_totals(self, daily_sales: List[Sale]) -> dict:
        """
        Calcula el total de productos y el importe total para un conjunto de ventas diarias.

        :param daily_sales: Lista de objetos Sale que contienen las ventas diarias.
        :return: Un diccionario con los totales mensuales:
                - total_products: Total de productos vendidos en el período.
                - total_income: Total de ingresos generados en el período.
        """
        total_products = 0
        total_income = 0.0

        for sale in daily_sales:
            # Sumar las cantidades y los ingresos de cada venta
            total_products += sum(detail.quantity for detail in sale.products)
            total_income += sale.total_amount

        return {
            "total_products": total_products,
            "total_income": round(
                total_income, 2
            ),  # Redondear a 2 decimales para moneda
        }

    def get_monthly_sales_by_product_data(
        self,
        month_str,
        business_id,
        specific_business_id=None,
        category_filter=None,
    ):
        """Genera la estructura diaria de ventas por producto para reportes mensuales."""
        sales, _ = self.get_daily_sales(
            month_str,
            business_id,
            specific_business_id,
        )

        sales_by_day = {}
        for sale in sales:
            date_key = sale.date.strftime("%Y-%m-%d")
            if date_key not in sales_by_day:
                sales_by_day[date_key] = {
                    "total_products": 0,
                    "total_income": 0,
                    "products": {},
                }

            sorted_products = sorted(
                sale.products,
                key=lambda sale_detail: sale_detail.product.name,
            )

            for sale_detail in sorted_products:
                product = sale_detail.product
                if category_filter and product.category != category_filter:
                    continue

                product_key = product.name
                if product_key not in sales_by_day[date_key]["products"]:
                    sales_by_day[date_key]["products"][product_key] = {
                        "quantity": 0,
                        "total_amount": 0,
                        "orders": set(),
                    }

                sales_by_day[date_key]["products"][product_key][
                    "quantity"
                ] += sale_detail.quantity
                if category_filter:
                    sales_by_day[date_key]["products"][product_key]["total_amount"] += (
                        sale_detail.quantity * product.price
                    )
                else:
                    sales_by_day[date_key]["products"][product_key][
                        "total_amount"
                    ] += sale_detail.total_price

                sales_by_day[date_key]["products"][product_key]["orders"].add(
                    (sale.id, sale.sale_number)
                )

                sales_by_day[date_key]["total_products"] += sale_detail.quantity
                if category_filter:
                    sales_by_day[date_key]["total_income"] += (
                        sale_detail.quantity * product.price
                    )
                else:
                    sales_by_day[date_key]["total_income"] += sale_detail.total_price

        daily_sales = []
        for date, data in sales_by_day.items():
            sorted_products = sorted(data["products"].items(), key=lambda item: item[0])
            daily_sales.append(
                {
                    "date": date,
                    "total_products": data["total_products"],
                    "total_income": data["total_income"],
                    "products": [
                        {
                            "name": name,
                            "quantity": product["quantity"],
                            "total_amount": product["total_amount"],
                            "orders": sorted(
                                list(product["orders"]),
                                key=lambda order: order[1],
                            ),
                        }
                        for name, product in sorted_products
                    ],
                }
            )

        return daily_sales

    @staticmethod
    def generate_excel_sales_by_product(data):
        """Genera un archivo Excel consolidado por producto para el período."""
        product_summary = defaultdict(
            lambda: {"quantity": 0, "total_amount": 0, "orders": defaultdict(set)}
        )

        for day in data:
            date_str = day["date"]
            day_number = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d")
            for product in day["products"]:
                product_name = product["name"]
                product_summary[product_name]["quantity"] += product["quantity"]
                product_summary[product_name]["total_amount"] += product["total_amount"]
                for order in product["orders"]:
                    product_summary[product_name]["orders"][day_number].add(order[1])

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte Mensual"
        sheet.append(["PRODUCTO", "CANTIDAD", "IMPORTE", "ORDENES"])

        for product_name, details in product_summary.items():
            orders_formatted = []
            for day_number, orders_set in details["orders"].items():
                orders_list = sorted(list(orders_set))
                orders_formatted.append(
                    f"{day_number}[{','.join(map(str, orders_list))}]"
                )

            sheet.append(
                [
                    product_name,
                    details["quantity"],
                    details["total_amount"],
                    ", ".join(orders_formatted),
                ]
            )

        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        return excel_file

    @staticmethod
    def generate_excel_sales_by_product_by_date(data):
        """Genera un archivo Excel desglosado por fecha y producto."""
        sales_by_day = defaultdict(list)
        for day in data:
            formatted_date = datetime.strptime(day["date"], "%Y-%m-%d").strftime(
                "%d/%m/%y"
            )
            for product in day["products"]:
                quantity = product["quantity"]
                total_amount = product["total_amount"]
                price_per_unit = total_amount / quantity if quantity > 0 else 0
                sales_by_day[formatted_date].append(
                    {
                        "product": product["name"],
                        "quantity": quantity,
                        "price_per_unit": price_per_unit,
                        "total_amount": total_amount,
                    }
                )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte Mensual"

        row_offset = 1
        for date, products in sales_by_day.items():
            sheet.cell(row=row_offset, column=1, value="DIA")
            sheet.cell(row=row_offset, column=2, value=date)
            row_offset += 1

            headers = ["PRODUCTO", "CANTIDAD", "PRECIO UNITARIO", "IMPORTE"]
            for col_idx, header in enumerate(headers, start=1):
                sheet.cell(row=row_offset, column=col_idx, value=header)
            row_offset += 1

            for product in products:
                sheet.cell(row=row_offset, column=1, value=product["product"])
                sheet.cell(row=row_offset, column=2, value=product["quantity"])
                sheet.cell(
                    row=row_offset,
                    column=3,
                    value=round(product["price_per_unit"], 2),
                )
                sheet.cell(
                    row=row_offset,
                    column=4,
                    value=round(product["total_amount"], 2),
                )
                row_offset += 1

            row_offset += 1

        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        return excel_file

    def get_ipv_daily_sales(self, month_str, business, business_filters):
        """Construye reporte IPV diario separando no-comida y comida."""
        selected_month, _, _ = self._month_date_range(month_str)
        sales = self._get_sales_for_month_by_scope(month_str, business_filters)

        non_food_products = {}
        for sale in sales:
            for sale_product in sale.products:
                product = sale_product.product
                if product.category not in ["comida", "postre", "trago", "cocteleria"]:
                    if product.name not in non_food_products:
                        non_food_products[product.name] = product.price

        sales_by_day = defaultdict(
            lambda: {
                "non_food": {
                    name: {
                        "quantity": 0,
                        "unit_price": price,
                        "total_price": 0,
                        "orders": set(),
                    }
                    for name, price in non_food_products.items()
                },
                "food": defaultdict(
                    lambda: {
                        "quantity": 0,
                        "unit_price": 0,
                        "total_price": 0,
                        "orders": set(),
                    }
                ),
            }
        )

        for sale in sales:
            date_key = sale.date.strftime("%d-%m-%Y")
            for sale_product in sale.products:
                product = sale_product.product
                if product.category == "cocteleria":
                    continue

                category_key = (
                    "food"
                    if product.category in ["comida", "postre", "trago"]
                    else "non_food"
                )
                product_entry = sales_by_day[date_key][category_key][product.name]
                product_entry["quantity"] += sale_product.quantity
                if sale_product.unit_price:
                    product_entry["unit_price"] = sale_product.unit_price
                product_entry["total_price"] += sale_product.total_price
                product_entry["orders"].add(sale.sale_number)

        daily_sales = []
        for date, categories in sales_by_day.items():
            non_food_list = [
                {
                    "name": name,
                    "quantity": data["quantity"],
                    "unit_price": round(data["unit_price"], 2),
                    "total_price": round(data["total_price"], 2),
                    "orders": sorted(data["orders"]),
                }
                for name, data in categories["non_food"].items()
            ]
            non_food_list.sort(key=lambda item: item["name"])

            food_list = [
                {
                    "name": name,
                    "quantity": data["quantity"],
                    "unit_price": round(data["unit_price"], 2),
                    "total_price": round(data["total_price"], 2),
                    "orders": sorted(data["orders"]),
                }
                for name, data in categories["food"].items()
            ]
            food_list.sort(key=lambda item: item["name"])

            daily_sales.append(
                {
                    "date": date,
                    "non_food": non_food_list,
                    "food": food_list,
                }
            )

        return daily_sales, selected_month

    # Helper methods

    def _group_sales_by_day(self, sales):
        """Agrupa las ventas por día."""
        sales_by_day = defaultdict(list)
        for sale in sales:
            date_key = sale.date.strftime("%Y-%m-%d")
            sales_by_day[date_key].append(sale)
        return sales_by_day

    def get_inventory_consumption(
        self, month_str: str, business_id: int, specific_business_id: int | None = None
    ) -> dict:
        """
        Calcula el consumo de materias primas (InventoryItem) derivado de las ventas de productos
        durante el mes indicado. Utiliza la relación Product -> ProductDetail (raw_materials)
        para saber cuánta materia prima consume cada unidad vendida.

        Retorna un diccionario donde la clave es el id de InventoryItem y el valor contiene
        nombre, unidad, total_consumed (float) y optionally product_usages (mapping product->qty).
        """
        _, start_date, end_date = self._month_date_range(month_str)

        # Obtener ventas del repositorio
        sales = self.repository.get_sales_for_month(
            business_id, specific_business_id, start_date, end_date
        )

        # Estructuras de acumulación
        consumption = defaultdict(
            lambda: {
                "inventory_item_id": None,
                "name": "",
                "unit": "",
                "total_consumed": 0.0,
                "product_usages": defaultdict(float),
            }
        )

        # Para cada venta y cada detalle, multiplicar la cantidad vendida por la cantidad
        # de materia prima requerida por unidad (ProductDetail.quantity)
        for sale in sales:
            for sd in sale.products:
                product = sd.product
                qty_sold = sd.quantity
                # Cada product.raw_materials describe el uso de inventory_item por unidad de producto
                for pd in getattr(product, "raw_materials", []):
                    inv = pd.raw_material
                    if inv is None:
                        continue
                    inv_id = inv.id
                    used_amount = pd.quantity * qty_sold
                    entry = consumption[inv_id]
                    entry["inventory_item_id"] = inv_id
                    entry["name"] = inv.name
                    entry["unit"] = inv.unit
                    entry["total_consumed"] += float(used_amount)
                    entry["product_usages"][product.name] += float(used_amount)

        # Convertir defaultdicts a dicts y product_usages a dict
        result = {}
        for inv_id, data in consumption.items():
            data_copy = dict(data)
            # product_usages es un defaultdict -> convertir
            data_copy["product_usages"] = dict(data_copy["product_usages"])
            # Redondear total_consumed a 4 decimales
            data_copy["total_consumed"] = round(data_copy["total_consumed"], 4)
            result[inv_id] = data_copy

        return result

    def get_inventory_consumption_by_day(
        self, month_str: str, business_id: int, specific_business_id: int | None = None
    ) -> list:
        """
        Calcula el consumo de materias primas por cada día del mes.

        Devuelve una lista de días donde cada día es un dict:
            {"date": "YYYY-MM-DD", "items": [{inventory_item_id, name, unit, total_consumed, product_usages}]}

        Los días con consumo cero no se incluyen.
        """
        _, start_date, end_date = self._month_date_range(month_str)

        # Obtener ventas del repositorio
        sales = self.repository.get_sales_for_month(
            business_id, specific_business_id, start_date, end_date
        )

        # Estructura: day_str -> inv_id -> entry
        days = defaultdict(
            lambda: defaultdict(
                lambda: {
                    "inventory_item_id": None,
                    "name": "",
                    "unit": "",
                    "total_consumed": 0.0,
                    "product_usages": defaultdict(float),
                }
            )
        )

        for sale in sales:
            day_key = sale.date.strftime("%Y-%m-%d")
            for sd in sale.products:
                product = sd.product
                qty_sold = sd.quantity
                for pd in getattr(product, "raw_materials", []):
                    inv = pd.raw_material
                    if inv is None:
                        continue
                    inv_id = inv.id
                    used_amount = pd.quantity * qty_sold
                    entry = days[day_key][inv_id]
                    entry["inventory_item_id"] = inv_id
                    entry["name"] = inv.name
                    entry["unit"] = inv.unit
                    entry["total_consumed"] += float(used_amount)
                    entry["product_usages"][product.name] += float(used_amount)

        # Convertir a lista ordenada por fecha
        result = []
        for day, items_map in sorted(days.items()):
            items_list = []
            for inv_id, data in items_map.items():
                data_copy = dict(data)
                data_copy["product_usages"] = dict(data_copy["product_usages"])
                data_copy["total_consumed"] = round(data_copy["total_consumed"], 4)
                items_list.append(data_copy)

            # Ordenar items por nombre
            items_list.sort(key=lambda x: x.get("name") or "")
            result.append({"date": day, "items": items_list})

        return result
